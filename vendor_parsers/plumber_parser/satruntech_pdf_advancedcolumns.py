import re
import sys
import io
from decimal import Decimal, InvalidOperation
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill
from datetime import datetime
import pdfplumber


sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

class EnhancedZohoInvoiceParser:
    def __init__(self):
        # Zoho Books field mappings
        self.zoho_header_fields = {
            'Bill Date': '',
            'Bill Number': '',
            'PurchaseOrder': '',
            'Bill Status': 'Open',
            'Source of Supply': '',
            'Destination of Supply': '',
            'GST Treatment': 'business_gst',
            'GST Identification Number (GSTIN)': '',
            'Is Inclusive Tax': 'FALSE',
            'TDS Percentage': '0',
            'TDS Amount': '0.00',
            'TDS Section Code': '',
            'TDS Name': '',
            'Vendor Name': '',
            'Due Date': '',
            'Currency Code': 'INR',
            'Exchange Rate': '1.00',
            'Attachment ID': '',
            'Attachment Preview ID': '',
            'Attachment Name': '',
            'Attachment Type': '',
            'Attachment Size': '',
            'SubTotal': '',
            'Total': '',
            'Balance': '',
            'Vendor Notes': '',
            'Terms & Conditions': '',
            'Payment Terms': '',
            'Payment Terms Label': '',
            'Is Billable': 'TRUE',
            'Customer Name': '',
            'Project Name': '',
            'Purchase Order Number': '',
            'Is Discount Before Tax': 'FALSE',
            'Entity Discount Amount': '0.00',
            'Discount Account': '',
            'Is Landed Cost': 'FALSE',
            'Warehouse Name': '',
            'Branch Name': '',
            'CF.Transporte_Name': '',
            'TCS Tax Name': '',
            'TCS Percentage': '0',
            'Nature Of Collection': '',
            'TCS Amount': '0.00',
            'Supply Type': 'Goods'
        }

        # Item level fields
        self.item_fields = {
            'Item Name': '',
            'SKU': '',
            'Item Description': '',
            'Account': 'Cost of Goods Sold',
            'Usage unit': 'Nos',
            'Quantity': '',
            'Rate': '',
            'Adjustment': '0.00',
            'Item Type': 'goods',
            'Tax Name': '',
            'Tax Percentage': '',
            'Tax Amount': '',
            'Tax Type': 'GST',
            'Item Exemption Code': '',
            'Reverse Charge Tax Name': '',
            'Reverse Charge Tax Rate': '0',
            'Reverse Charge Tax Type': '',
            'Item Total': '',
            'HSN/SAC': '',
            'ITC Eligibility': 'Eligible'
        }

   
    
    def safe_convert(self, value, default=Decimal("0.0")):
        try:
            return Decimal(str(value).replace(",", "").strip())
        except (InvalidOperation, ValueError, TypeError):
            return default

    def safe_decimal(self, value):
        try:
            return Decimal(str(value).replace(',', '').strip())
        except:
            return Decimal("0.00")

    def clean_amount(self, value):
        """Clean unwanted characters and safely convert to numeric string"""
        return re.sub(r'[^\d.]', '', value)

    def extract_items_from_text(self, text):
        lines = text.splitlines()
        items = []

        for line in lines:
            if len(line.strip()) < 20:
                continue

            clean = re.sub(r'[\[\]\|_}{)(]', '', line).strip()
            prices = re.findall(r'[\d,]+\.\d{2}', clean)
            if len(prices) < 5:
                continue

            hsn_match = re.search(r'\b(\d{4,8})\b', clean)
            if not hsn_match:
                continue
            hsn_code = hsn_match.group(1)

            tax_percent_match = re.search(r'(\d{1,2})%', clean)
            tax_percent = tax_percent_match.group(1) if tax_percent_match else "0"

            start_match = re.match(r'\s*(\d+)[\)/]?\s*(.+?)\b' + re.escape(hsn_code) + r'\b\s+(\d+)', clean)
            if not start_match:
                continue

            item_no = start_match.group(1)
            description = start_match.group(2).strip()
            qty = start_match.group(3)

            # Price field positions based on your format
            rate = self.safe_decimal(prices[-5])
            taxable_value = self.safe_decimal(prices[-4])
            tax_amount = self.safe_decimal(prices[-2])
            total = self.safe_decimal(prices[-1])

            items.append({
                "item_no": item_no,
                "description": description if description != "/" else "Unknown Item",
                "hsn_code": hsn_code,
                "qty": self.safe_decimal(qty),
                "rate": rate,
                "taxable_value": taxable_value,
                "tax_percent": tax_percent,
                "tax_amount": tax_amount,
                "total": total
            })

        return items

    def extract_invoice_data(self, text):
        invoice = {}
        
        # Invoice number
        match = re.search(r'invoice No:\s*(\S+)', text, re.IGNORECASE)
        if match:
            invoice['invoice_number'] = match.group(1)
        
        # Invoice date
        match = re.search(r'invoice date:\s*(\d{2}/\d{2}/\d{4})', text, re.IGNORECASE)
        if match:
            invoice['invoice_date'] = match.group(1)
        
        # Transport mode
        match = re.search(r'Transport Mode:\s*([^\n]+)', text)
        if match:
            invoice['transport_mode'] = match.group(1).strip()
        
        # State and code
        match = re.search(r'State:\s*(.+?)\s+Code\s+(\d+)', text)
        if match:
            invoice['state'] = match.group(1).strip()
            invoice['state_code'] = match.group(2)

        return invoice

    def extract_seller_buyer(self, text):
        # Seller information
        seller = {"company_name": "SATRUN TECHNOLOGIES"}
        
        match = re.search(r'Mobile:([\d/]+)', text)
        if match:
            seller["phone"] = match.group(1)
        
        match = re.search(r'Email:\s*([^\s\n]+)', text)
        if match:
            seller["email"] = match.group(1)
        
        match = re.search(r'GSTIN:\s*(\S+)', text)
        if match:
            seller["gstin"] = match.group(1)

        # Buyer information
        buyer = {}
        match = re.search(r'Bill to Party\s*\|?(.+?)\n', text)
        if match:
            buyer["company_name"] = match.group(1).strip()
        
        match = re.search(r'IGST NO\s*:\s*(\S+)', text)
        if match:
            buyer["gstin"] = match.group(1)
        
        match = re.search(r'State Name:\s*(.+?),', text)
        if match:
            buyer["state"] = match.group(1).strip()

        return seller, buyer

    def extract_amount_and_totals(self, text):
        amount = {}
        
        # Amount in words
        match = re.search(r'INR\s+([A-Z\s\-]+ONLY)', text, re.IGNORECASE)
        if match:
            amount["amount_in_words"] = "INR " + match.group(1).strip()
        else:
            amount["amount_in_words"] = "INR ZERO ONLY"
        
        # Total amount after tax
        match = re.search(r'Total Amount after Tax\s+([\d,]+\.\d{2})', text)
        if match:
            amount["total_amount"] = match.group(1)
        else:
            amount["total_amount"] = "0.00"

        # Subtotal (amount before tax)
        match = re.search(r'Total Amount before Tax\s+([\d,]+\.\d{2})', text)
        if match:
            amount["subtotal"] = match.group(1)
        else:
            amount["subtotal"] = "0.00"

        return amount

    def extract_taxes(self, text):
        taxes = {}
        text = text.replace('\n', ' ').replace('\r', ' ')
        text = re.sub(r'[\[\]{}|()_]', '', text)

        # Total tax amount
        total_match = re.search(r'Total\s+Tax\s+Amount[.:]?\s*([\d,]+\.\d{2})', text, re.IGNORECASE)
        if total_match:
            taxes["total_tax"] = self.clean_amount(total_match.group(1))

        # IGST percentage
        igst_match = re.search(r'IGST@(\d{1,2})%', text)
        if igst_match:
            taxes["igst_percent"] = igst_match.group(1)

        return taxes

    def map_to_zoho_format(self, invoice_meta, seller, buyer, items_data, amount, taxes):
        """Map extracted data to Zoho Books format"""
        zoho_data = self.zoho_header_fields.copy()
        
        # Map header fields
        zoho_data['Bill Date'] = invoice_meta.get('invoice_date', '')
        zoho_data['Bill Number'] = invoice_meta.get('invoice_number', '')
        zoho_data['Vendor Name'] = seller.get('company_name', '')
        zoho_data['Customer Name'] = buyer.get('company_name', '')
        zoho_data['GST Identification Number (GSTIN)'] = buyer.get('gstin', '')
        zoho_data['Source of Supply'] = invoice_meta.get('state_code', '')
        zoho_data['Destination of Supply'] = buyer.get('state_code', '')
        zoho_data['SubTotal'] = amount.get('subtotal', '0.00').replace(',', '')
        zoho_data['Total'] = amount.get('total_amount', '0.00').replace(',', '')
        zoho_data['Balance'] = amount.get('total_amount', '0.00').replace(',', '')
        zoho_data['Payment Terms'] = 'Due on Receipt'
        zoho_data['Due Date'] = invoice_meta.get('invoice_date', '')
        
        # Determine tax type
        if taxes.get('igst_percent'):
            zoho_data['Tax Name'] = f"IGST{taxes['igst_percent']}"
            zoho_data['Tax Percentage'] = taxes['igst_percent']
            zoho_data['Tax Amount'] = taxes.get('total_tax', '0.00')

        # Map items
        zoho_items = []
        for item in items_data:
            zoho_item = self.item_fields.copy()
            zoho_item.update({
                'Item Name': item['description'][:50],  # Truncate for name
                'SKU': item['hsn_code'],
                'Item Description': item['description'],
                'HSN/SAC': item['hsn_code'],
                'Quantity': str(item['qty']),
                'Rate': str(item['rate']).replace(',', ''),
                'Item Total': str(item['total']).replace(',', ''),
                'Tax Percentage': item['tax_percent'],
                'Tax Amount': str(item['tax_amount']).replace(',', ''),
                'Tax Name': f"IGST{item['tax_percent']}" if item['tax_percent'] else ""
            })
            zoho_items.append(zoho_item)

        return {
            'invoice_header': zoho_data,
            'invoice_items': zoho_items,
            'summary': {
                'total_items': len(zoho_items),
                'total_quantity': sum(float(item['Quantity']) for item in zoho_items),
                'subtotal': zoho_data['SubTotal'],
                'total': zoho_data['Total']
            }
        }

    def create_excel_export(self, zoho_data, output_filename):
        """Create single-sheet Excel file in Zoho Books import format"""
        try:
            # All Zoho columns in correct order
            zoho_columns = [
                'Bill Date', 'Bill Number', 'PurchaseOrder', 'Bill Status', 'Source of Supply',
                'Destination of Supply', 'GST Treatment', 'GST Identification Number (GSTIN)',
                'Is Inclusive Tax', 'TDS Percentage', 'TDS Amount', 'TDS Section Code', 'TDS Name',
                'Vendor Name', 'Due Date', 'Currency Code', 'Exchange Rate', 'Attachment ID',
                'Attachment Preview ID', 'Attachment Name', 'Attachment Type', 'Attachment Size',
                'Item Name', 'SKU', 'Item Description', 'Account', 'Usage unit', 'Quantity',
                'Rate', 'Adjustment', 'Item Type', 'Tax Name', 'Tax Percentage', 'Tax Amount',
                'Tax Type', 'Item Exemption Code', 'Reverse Charge Tax Name',
                'Reverse Charge Tax Rate', 'Reverse Charge Tax Type', 'Item Total', 'SubTotal',
                'Total', 'Balance', 'Vendor Notes', 'Terms & Conditions', 'Payment Terms',
                'Payment Terms Label', 'Is Billable', 'Customer Name', 'Project Name',
                'Purchase Order Number', 'Is Discount Before Tax', 'Entity Discount Amount',
                'Discount Account', 'Is Landed Cost', 'Warehouse Name', 'Branch Name',
                'CF.Transporte_Name', 'TCS Tax Name', 'TCS Percentage', 'Nature Of Collection',
                'TCS Amount', 'HSN/SAC', 'Supply Type', 'ITC Eligibility'
            ]

            # Create workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            if ws is not None:
                ws.title = "Zoho Books Import"
            else:
                ws = wb.create_sheet("Zoho Books Import")

            # Add header row with styling
            ws.append(zoho_columns)
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")

            for col_num in range(1, len(zoho_columns) + 1):
                cell = ws.cell(row=1, column=col_num)
                cell.font = header_font
                cell.fill = header_fill

            # Get data
            header_data = zoho_data['invoice_header']
            items_data = zoho_data['invoice_items']

            # Add data rows
            if not items_data:
                # No items - create single row with header data
                row_data = []
                for col in zoho_columns:
                    row_data.append(str(header_data.get(col, '')))
                ws.append(row_data)
            else:
                # Add one row per item
                for item in items_data:
                    row_data = []
                    for col in zoho_columns:
                        if col in item and item[col]:
                            row_data.append(str(item[col]))
                        elif col in header_data:
                            row_data.append(str(header_data[col]))
                        else:
                            row_data.append('')
                    ws.append(row_data)

            # Auto-adjust column widths
            for col_num, column in enumerate(zoho_columns, 1):
                column_letter = openpyxl.utils.get_column_letter(col_num)
                if column in ['Item Description', 'Item Name', 'Vendor Name']:
                    ws.column_dimensions[column_letter].width = 30
                elif column in ['Bill Number', 'GST Identification Number (GSTIN)']:
                    ws.column_dimensions[column_letter].width = 20
                else:
                    ws.column_dimensions[column_letter].width = 15

            wb.save(output_filename)
            return True

        except Exception as e:
            print(f"Error creating Excel: {e}")
            return False

    def process_invoice(self, text):
        """Main processing function that returns Zoho-formatted data"""
        try:
            # Extract all data using your existing methods
                   # Extract all needed info from text via your own extraction functions:
            invoice_meta = self.extract_invoice_data(text)
            seller_info, buyer_info = self.extract_seller_buyer(text)
            items_raw = self.extract_items_from_text(text)
            totals = self.extract_amount_and_totals(text)
            # Add other extraction calls as needed but no change to your internal logic

            # Prepare invoice_data dict with all required Zoho fields, set empty string if unavailable
            invoice_data = {
                'Bill Date': invoice_meta.get('invoice_date', ''),
                'Bill Number': invoice_meta.get('invoice_number', ''),
                'PurchaseOrder': '',  # no extractor? Leave empty
                'Bill Status': 'Open',
                'Source of Supply': '',  # no extractor? Leave empty
                'Destination of Supply': '',  # no extractor? Leave empty
                'GST Treatment': 'business_gst',
                'GST Identification Number (GSTIN)': seller_info.get('gstin', ''),
                'Is Inclusive Tax': 'FALSE',
                'TDS Percentage': '0',
                'TDS Amount': '0.00',
                'TDS Section Code': '',
                'TDS Name': '',
                'Vendor Name': seller_info.get('company_name', ''),
                'Due Date': '',  # no extractor? Leave empty
                'Currency Code': 'INR',
                'Exchange Rate': '1.00',
                'Attachment ID': '',
                'Attachment Preview ID': '',
                'Attachment Name': '',
                'Attachment Type': '',
                'Attachment Size': '',
                'SubTotal': totals.get('subtotal', '0.00'),
                'Total': totals.get('total_amount', '0.00'),
                'Balance': totals.get('total_amount', '0.00'),
                'Vendor Notes': '',
                'Terms & Conditions': '',
                'Payment Terms': '',
                'Payment Terms Label': '',
                'Is Billable': 'TRUE',
                'Customer Name': buyer_info.get('company_name', ''),
                'Project Name': '',
                'Purchase Order Number': '',
                'Is Discount Before Tax': 'FALSE',
                'Entity Discount Amount': '0.00',
                'Discount Account': '',
                'Is Landed Cost': 'FALSE',
                'Warehouse Name': '',
                'Branch Name': '',
                'CF.Transporte_Name': invoice_meta.get('transport_mode', ''),
                'TCS Tax Name': '',
                'TCS Percentage': '0',
                'Nature Of Collection': '',
                'TCS Amount': '0.00',
                'Supply Type': 'Goods',
                'ITC Eligibility': 'Eligible'  # If header level not present, can keep default
            }

            # Prepare items list with all required Zoho item fields, safely extract or set ""
            items = []
            for raw in items_raw:
                item = {
                    'Item Name': raw.get('description', ''),
                    'SKU': raw.get('hsn_code', ''),
                    'Item Description': raw.get('description', ''),
                    'Account': 'Cost of Goods Sold',
                    'Usage unit': 'Nos',
                    'Quantity': str(raw.get('qty', '')),
                    'Rate': str(raw.get('rate', '')),
                    'Adjustment': '0.00',
                    'Item Type': 'goods',
                    'Tax Name': '',  # Not provided in enhanced parser, leave empty
                    'Tax Percentage': str(raw.get('tax_percent', '0')),
                    'Tax Amount': str(raw.get('tax_amount', '0.00')),
                    'Tax Type': 'GST',
                    'Item Exemption Code': '',
                    'Reverse Charge Tax Name': '',
                    'Reverse Charge Tax Rate': '0',
                    'Reverse Charge Tax Type': '',
                    'Item Total': str(raw.get('total', '0.00')),
                    'HSN/SAC': raw.get('hsn_code', ''),
                    'Supply Type': 'Goods',
                    'ITC Eligibility': 'Eligible'
                }
                items.append(item)

            # Return in expected format
            return {
                "invoice_data": invoice_data,
                "items": items
            }


        except Exception as e:
            print(f"Error processing invoice: {e}")
            import traceback
            traceback.print_exc()
            return None

# Usage function that maintains your existing interface
def process_invoice(text, path):
    """Wrapper function to maintain compatibility with your existing code"""
    parser = EnhancedZohoInvoiceParser()
    text = text
    unwanted_path = path
    result = parser.process_invoice(text)
    
    if result:
        # Display summary
        print("\n" + "="*60)
        print("ZOHO BOOKS MAPPING COMPLETE")
        print("="*60)
        print(f"Invoice Number: {result['invoice_data']['Bill Number']}")
        print(f"Vendor: {result['invoice_data']['Vendor Name']}")
        print(f"Customer: {result['invoice_data']['Customer Name']}")
        print(f"Total Amount: ₹{result['invoice_data']['Total']}")
        print(f"GSTIN: {result['invoice_data']['GST Identification Number (GSTIN)']}")
        
        if 'excel_file' in result:
            print(f"\nExcel Export: {result['excel_file']}")
        print("="*60)
        
        return result
    else:
        return {
            "invoice_data": {},
            "items": []
        }
    
# def extract_text(path):
#         with pdfplumber.open(path) as pdf:
#             return "\n".join([p.extract_text() for p in pdf.pages])
        

# # Example usage
# if __name__ == "__main__":
#     invoice_path = r"uploads\satrunInvoicetext.pdf"
#     text = extract_text(invoice_path)

#     if not text.strip():
#         raise ValueError("[ERROR] No text found in PDF. It's likely scanned. Use OCR mode.")
#     result = process_invoice(text, invoice_path)
#     print("\nProcessing complete!")