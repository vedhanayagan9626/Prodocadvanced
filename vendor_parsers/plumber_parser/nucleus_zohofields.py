# -*- coding: utf-8 -*-
import pdfplumber
import re
from paddleocr import PaddleOCR
import numpy as np
from decimal import Decimal
from datetime import datetime
import logging
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import pandas as pd
import os

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

class ZohoInvoiceParser:
    def __init__(self):
        self.ocr = PaddleOCR(use_angle_cls=True, lang='en')
        self.items = []
        self.tax_info = {}
        self.grand_total = {}
        self.current_item = None
        self.capture_description = False
        
        # Zoho Books field mappings with proper order
        self.zoho_header_fields = [
            'Bill Date', 'Bill Number', 'Purchase Order', 'Bill Status', 
            'Source of Supply', 'Destination of Supply', 'GST Treatment', 
            'GST Identification Number (GSTIN)', 'Is Inclusive Tax', 'TDS Percentage', 
            'TDS Amount', 'TDS Section Code', 'TDS Name', 'Vendor Name', 
            'Due Date', 'Currency Code', 'Exchange Rate', 'Attachment ID', 
            'Attachment Preview ID', 'Attachment Name', 'Attachment Type', 
            'Attachment Size', 'SubTotal', 'Total', 'Balance', 'Vendor Notes', 
            'Terms & Conditions', 'Payment Terms', 'Payment Terms Label', 
            'Is Billable', 'Customer Name', 'Project Name', 'Purchase Order Number', 
            'Is Discount Before Tax', 'Entity Discount Amount', 'Discount Account', 
            'Is Landed Cost', 'Warehouse Name', 'Branch Name', 'CF.Transporte_Name', 
            'TCS Tax Name', 'TCS Percentage', 'Nature Of Collection', 'TCS Amount', 
            'Supply Type'
        ]
        
        # Item-level fields with proper order
        self.zoho_item_fields = [
            'Item Name', 'SKU', 'Item Description', 'Account', 'Usage unit', 
            'Quantity', 'Rate', 'Adjustment', 'Item Type', 'Tax Name', 
            'Tax Percentage', 'Tax Amount', 'Tax Type', 'Item Exemption Code', 
            'Reverse Charge Tax Name', 'Reverse Charge Tax Rate', 
            'Reverse Charge Tax Type', 'Item Total', 'HSN/SAC', 'ITC Eligibility'
        ]
        
        self.zoho_fields = {
            # Header fields
            'Bill Date': '',
            'Bill Number': '',
            'Purchase Order': '',
            'Bill Status': 'Open',  # Default status
            'Source of Supply': '',
            'Destination of Supply': '',
            'GST Treatment': 'GST Registered',  # Default for Indian invoices
            'GST Identification Number (GSTIN)': '',
            'Is Inclusive Tax': 'No',  # Default
            'TDS Percentage': '0',
            'TDS Amount': '0.00',
            'TDS Section Code': '',
            'TDS Name': '',
            'Vendor Name': '',
            'Due Date': '',
            'Currency Code': 'INR',  # Default for Indian invoices
            'Exchange Rate': '1.00',  # Default for INR
            'Attachment ID': '',
            'Attachment Preview ID': '',
            'Attachment Name': '',
            'Attachment Type': '',
            'Attachment Size': '',
            
            # Totals
            'SubTotal': '',
            'Total': '',
            'Balance': '',
            'Vendor Notes': '',
            'Terms & Conditions': '',
            'Payment Terms': '',
            'Payment Terms Label': '',
            'Is Billable': 'Yes',
            'Customer Name': '',
            'Project Name': '',
            'Purchase Order Number': '',
            'Is Discount Before Tax': 'No',
            'Entity Discount Amount': '0.00',
            'Discount Account': '',
            'Is Landed Cost': 'No',
            'Warehouse Name': '',
            'Branch Name': '',
            'CF.Transporte_Name': '',
            'TCS Tax Name': '',
            'TCS Percentage': '0',
            'Nature Of Collection': '',
            'TCS Amount': '0.00',
            'Supply Type': 'Goods'  # Default
        }
        
        # Item-level fields template
        self.item_template = {
            'Item Name': '',
            'SKU': '',
            'Item Description': '',
            'Account': '',
            'Usage unit': '',
            'Quantity': '',
            'Rate': '',
            'Adjustment': '0.00',
            'Item Type': 'Goods',
            'Tax Name': '',
            'Tax Percentage': '',
            'Tax Amount': '',
            'Tax Type': 'GST',
            'Item Exemption Code': '',
            'Reverse Charge Tax Name': '',
            'Reverse Charge Tax Rate': '0',
            'Reverse Charge Tax Type': '',
            'Item Total': '',
            'HSN/SAC': '',
            'ITC Eligibility': 'Eligible'  # Default
        }

    def clean_number(self, value):
        """Clean numeric values by removing commas and handling decimals"""
        if isinstance(value, str):
            # Remove commas and clean up formatting
            cleaned = value.replace(",", "").replace("₹", "").strip()
            # Handle cases where there might be multiple dots
            if cleaned.count(".") > 1:
                parts = cleaned.split(".")
                cleaned = "".join(parts[:-1]) + "." + parts[-1]
            return cleaned
        return str(value) if value is not None else "0"

    def safe_convert(self, value, to_type):
        """Safely convert value to specified type"""
        try:
            if to_type == float:
                return float(self.clean_number(value))
            elif to_type == int:
                return int(float(self.clean_number(value)))
            elif to_type == Decimal:
                return Decimal(self.clean_number(value))
            elif to_type == str:
                return str(value).strip() if value else ""
            else:
                return value
        except (ValueError, TypeError, Exception):
            return 0 if to_type in [float, int, Decimal] else ""

    def format_currency_indian(self, value):
        """Format currency in Indian numbering system"""
        try:
            amount = float(self.clean_number(value))
        except:
            return "0.00"
        
        # Format to 2 decimal places
        formatted = f"{amount:.2f}"
        parts = formatted.split(".")
        integer = parts[0]
        decimal = parts[1]
        
        # Apply Indian numbering (lakhs, crores)
        if len(integer) <= 3:
            result = integer
        else:
            result = integer[-3:]
            integer = integer[:-3]
            while len(integer) > 0:
                if len(integer) > 2:
                    result = integer[-2:] + "," + result
                    integer = integer[:-2]
                else:
                    result = integer + "," + result
                    break
        
        return result + "." + decimal

    def extract_invoice_metadata(self, lines):
        """Extract invoice header information"""
        metadata = {}
        
        for i, line in enumerate(lines):
            line = line.strip()
            
            # Invoice Number
            if not metadata.get("invoice_number"):
                inv_match = re.search(r"Invoice\s*No[:\-]?\s*(\S+)", line, re.IGNORECASE)
                if inv_match:
                    metadata["invoice_number"] = inv_match.group(1).strip()
            
            # Invoice Date
            if not metadata.get("invoice_date"):
                date_patterns = [
                    r"Invoice\s*Date[:\-]?\s*([0-9]{1,2}[\/\-\s]?[A-Za-z]{3,9}[\/\-\s]?[0-9]{2,4})",
                    r"Date[:\-]?\s*([0-9]{1,2}[\/\-\s]?[A-Za-z]{3,9}[\/\-\s]?[0-9]{2,4})",
                    r"([0-9]{1,2}\s+[A-Za-z]{3,9}\s+[0-9]{4})"
                ]
                for pattern in date_patterns:
                    date_match = re.search(pattern, line, re.IGNORECASE)
                    if date_match:
                        metadata["invoice_date"] = date_match.group(1).strip()
                        break
            
            # Vendor/Seller Information
            if "nucleus analytics" in line.lower():
                metadata["vendor_name"] = "Nucleus Analytics Private Limited"
            
            # GSTIN extraction
            gstin_match = re.search(r"GSTIN[:\s]*([0-9A-Z]{15})", line, re.IGNORECASE)
            if gstin_match:
                metadata["vendor_gstin"] = gstin_match.group(1)
            
            # Customer GSTIN
            customer_gstin_match = re.search(r"Customer\s+GSTIN[\s_:]+([0-9A-Z]{15})", line, re.IGNORECASE)
            if customer_gstin_match:
                metadata["customer_gstin"] = customer_gstin_match.group(1)
            
            # Customer Name (M/S pattern)
            if "M/S" in line.upper() and not metadata.get("customer_name"):
                parties = re.findall(r"M/S[.,]?\s*(.*?)(?:\s{2,}|$)", line, re.IGNORECASE)
                for party in parties:
                    name = party.strip(" ,.")
                    if name:
                        metadata["customer_name"] = name
                        break
            
            # Payment Terms
            payment_match = re.search(r"Payment\s+terms[:\s]+([^|]+)", line, re.IGNORECASE)
            if payment_match:
                metadata["payment_terms"] = payment_match.group(1).strip()
            
            # Place of Supply
            supply_match = re.search(r"Place\s+of\s+Supply\s+state\s+code[:\s]+(\d+)", line, re.IGNORECASE)
            if supply_match:
                metadata["source_of_supply"] = supply_match.group(1)
            
            # Place of Delivery
            delivery_match = re.search(r"Place\s+of\s+Delivery\s+state\s+code[:\s]+(\d+)", line, re.IGNORECASE)
            if delivery_match:
                metadata["destination_of_supply"] = delivery_match.group(1)
        
        return metadata

    def extract_line_items(self, lines):
        """Extract line items from the invoice"""
        items = []
        current_item = None
        capture_description = False
        
        # Regex pattern for line items
        item_regex = re.compile(
            r"^(\d+)\s+(\d{8})\s+(.+?)\s+(Nos|PCS|Units?)\s*\|\s*([\d.,]+)\s+([\d.,]+)\s+([\d.,]+)"
        )
        
        for i, line in enumerate(lines):
            line = line.strip()
            
            # Check for item match
            item_match = item_regex.match(line)
            if item_match:
                # Save previous item if exists
                if current_item:
                    items.append(current_item)
                
                # Create new item
                quantity = self.safe_convert(item_match.group(5), float)
                rate = self.safe_convert(item_match.group(6), float)
                total = self.safe_convert(item_match.group(7), float)
                
                current_item = {
                    'Item Name': item_match.group(3).strip(),
                    'SKU': item_match.group(2),  # HSN Code as SKU
                    'Item Description': item_match.group(3).strip(),
                    'HSN/SAC': item_match.group(2),
                    'Usage unit': item_match.group(4),
                    'Quantity': str(quantity),
                    'Rate': str(rate),
                    'Item Total': str(total),
                    'Item Type': 'Goods',
                    'Tax Type': 'GST',
                    'ITC Eligibility': 'Eligible',
                    'Account': 'Cost of Goods Sold',  # Default account
                    'Adjustment': '0.00'
                }
                capture_description = True
                continue
            
            # Capture additional description lines
            if capture_description and current_item:
                if re.search(r"(Declaration|Subtotal|Total|GSTIN|Authorised|Bank|NEFT|Amount Due)", line, re.IGNORECASE):
                    capture_description = False
                else:
                    # Append to description
                    current_item['Item Description'] += " " + line.strip()
                    current_item['Item Name'] = current_item['Item Description'][:50]  # Truncate for name
        
        # Add final item
        if current_item:
            items.append(current_item)
        
        return items

    def extract_tax_information(self, lines):
        """Extract tax information from invoice"""
        tax_info = {}
        
        for line in lines:
            line = line.strip()
            
            # IGST
            igst_match = re.search(r"IGST\s*%?\s*(\d+)%?\s+([\d.,]+)", line, re.IGNORECASE)
            if igst_match:
                tax_info['igst_percentage'] = igst_match.group(1)
                tax_info['igst_amount'] = self.clean_number(igst_match.group(2))
            
            # CGST
            cgst_match = re.search(r"CGST\s*%?\s*(\d+)%", line, re.IGNORECASE)
            if cgst_match:
                tax_info['cgst_percentage'] = cgst_match.group(1)
            
            # SGST
            sgst_match = re.search(r"SGST\s*%?\s*(\d+)%", line, re.IGNORECASE)
            if sgst_match:
                tax_info['sgst_percentage'] = sgst_match.group(1)
        
        return tax_info

    def extract_totals(self, lines):
        """Extract total amounts from invoice"""
        totals = {}
        
        for line in lines:
            line = line.strip()
            
            # Subtotal
            sub_match = re.search(r"Sub[\s\-]?Total[:\s]+([\d,]+\.?\d*)", line, re.IGNORECASE)
            if sub_match:
                totals['subtotal'] = self.clean_number(sub_match.group(1))
            
            # Total
            total_match = re.search(r"(?:Total|Grand\s+Total)[:\s]+([\d,]+\.?\d*)", line, re.IGNORECASE)
            if total_match and 'total' not in totals:  # Avoid overwriting with subtotal
                totals['total'] = self.clean_number(total_match.group(1))
            
            # Amount in words
            words_match = re.search(r"Rupees\s+(.+?)\s+Only", line, re.IGNORECASE)
            if words_match:
                totals['amount_in_words'] = "Rupees " + words_match.group(1).strip() + " Only"
        
        return totals

    def calculate_due_date(self, invoice_date, payment_terms):
        """Calculate due date based on payment terms"""
        try:
            # Parse invoice date (assuming DD/MM/YYYY or DD-MMM-YYYY format)
            if not invoice_date:
                return ""
            
            # For "Against Invoice" or similar terms, due date = invoice date
            if "against invoice" in payment_terms.lower():
                return invoice_date
            
            # Extract number of days from payment terms
            days_match = re.search(r"(\d+)\s*days?", payment_terms, re.IGNORECASE)
            if days_match:
                days = int(days_match.group(1))
                # For simplicity, return the same date + days indication
                return f"{invoice_date} + {days} days"
            
            return invoice_date
        except:
            return invoice_date

    def map_to_zoho_format(self, metadata, items, tax_info, totals):
        """Map extracted data to Zoho Books format"""
        zoho_data = self.zoho_fields.copy()
        
        # Map header information
        zoho_data['Bill Date'] = metadata.get('invoice_date', '')
        zoho_data['Bill Number'] = metadata.get('invoice_number', '')
        zoho_data['Vendor Name'] = metadata.get('vendor_name', 'Nucleus Analytics Private Limited')
        zoho_data['GST Identification Number (GSTIN)'] = metadata.get('customer_gstin', '')
        zoho_data['Source of Supply'] = metadata.get('source_of_supply', '')
        zoho_data['Destination of Supply'] = metadata.get('destination_of_supply', '')
        zoho_data['Customer Name'] = metadata.get('customer_name', '')
        zoho_data['Payment Terms'] = metadata.get('payment_terms', '')
        zoho_data['Due Date'] = self.calculate_due_date(
            metadata.get('invoice_date', ''), 
            metadata.get('payment_terms', '')
        )
        
        # Map totals
        zoho_data['SubTotal'] = totals.get('subtotal', '0.00')
        zoho_data['Total'] = totals.get('total', '0.00')
        zoho_data['Balance'] = totals.get('total', '0.00')  # Assuming unpaid
        
        # Tax information
        if tax_info.get('igst_percentage'):
            zoho_data['Tax Name'] = 'IGST'
            zoho_data['Tax Percentage'] = tax_info['igst_percentage']
            zoho_data['Tax Amount'] = tax_info.get('igst_amount', '0.00')
        elif tax_info.get('cgst_percentage'):
            zoho_data['Tax Name'] = 'CGST+SGST'
            zoho_data['Tax Percentage'] = str(
                int(tax_info.get('cgst_percentage', '0')) + 
                int(tax_info.get('sgst_percentage', '0'))
            )
        
        # Map items
        zoho_items = []
        for item in items:
            zoho_item = self.item_template.copy()
            zoho_item.update(item)
            
            # Add tax information to items
            if tax_info.get('igst_percentage'):
                zoho_item['Tax Name'] = 'IGST'
                zoho_item['Tax Percentage'] = tax_info['igst_percentage']
            elif tax_info.get('cgst_percentage'):
                zoho_item['Tax Name'] = 'CGST+SGST'
                total_tax = int(tax_info.get('cgst_percentage', '0')) + int(tax_info.get('sgst_percentage', '0'))
                zoho_item['Tax Percentage'] = str(total_tax)
            
            # Calculate tax amount for item
            item_total = self.safe_convert(zoho_item['Item Total'], float)
            tax_percentage = self.safe_convert(zoho_item['Tax Percentage'], float)
            if int(item_total) > 0 and int(tax_percentage) > 0:
                # Assuming tax is calculated on item total
                tax_amount = (int(item_total) * int(tax_percentage)) / (100 + int(tax_percentage))
                zoho_item['Tax Amount'] = str(round(tax_amount, 2))
            
            zoho_items.append(zoho_item)
        
        return {
            'invoice_header': zoho_data,
            'invoice_items': zoho_items,
            'summary': {
                'total_items': len(zoho_items),
                'total_quantity': sum(float(self.safe_convert(item['Quantity'], float)) for item in zoho_items),
                'subtotal': zoho_data['SubTotal'],
                'total': zoho_data['Total']
            }
        }

    def export_to_excel(self, zoho_data, excel_filename=None):
        """Export the mapped data to Excel format compatible with Zoho Books"""
        try:
            if not excel_filename:
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                excel_filename = f"zoho_invoice_export_{timestamp}.xlsx"
            
            # Create a new workbook
            wb = Workbook()
            
            # Remove default worksheet
            if 'Sheet' in wb.sheetnames:
                std = wb['Sheet']
                wb.remove(std)
            
            # Create worksheet for Zoho Import Format
            ws_zoho = wb.create_sheet("Zoho_Import_Format")
            
            # Prepare data for Zoho format (each item is a separate row with header info repeated)
            all_rows = []
            header_data = zoho_data['invoice_header']
            
            for item in zoho_data['invoice_items']:
                row_data = []
                
                # Add header fields first
                for field in self.zoho_header_fields:
                    row_data.append(header_data.get(field, ''))
                
                # Add item fields
                for field in self.zoho_item_fields:
                    row_data.append(item.get(field, ''))
                
                all_rows.append(row_data)
            
            # Create combined header
            combined_header = self.zoho_header_fields + self.zoho_item_fields
            
            # Write header row
            for col_num, header in enumerate(combined_header, 1):
                cell = ws_zoho.cell(row=1, column=col_num, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
            
            # Write data rows
            for row_num, row_data in enumerate(all_rows, 2):
                for col_num, value in enumerate(row_data, 1):
                    ws_zoho.cell(row=row_num, column=col_num, value=value)
            
            # Auto-adjust column widths
            for column_cells in ws_zoho.columns:
                length = max(len(str(cell.value or '')) for cell in column_cells)
                ws_zoho.column_dimensions[column_cells[0].column_letter].width = min(length + 2, 50)
            
            # Create summary worksheet
            ws_summary = wb.create_sheet("Invoice_Summary")
            
            # Summary data
            summary_data = [
                ["Invoice Summary", ""],
                ["Invoice Number", header_data.get('Bill Number', '')],
                ["Invoice Date", header_data.get('Bill Date', '')],
                ["Vendor Name", header_data.get('Vendor Name', '')],
                ["Customer Name", header_data.get('Customer Name', '')],
                ["GSTIN", header_data.get('GST Identification Number (GSTIN)', '')],
                ["", ""],
                ["Financial Summary", ""],
                ["Total Items", zoho_data['summary']['total_items']],
                ["Total Quantity", zoho_data['summary']['total_quantity']],
                ["Subtotal", f"₹{zoho_data['summary']['subtotal']}"],
                ["Total", f"₹{zoho_data['summary']['total']}"],
                ["", ""],
                ["Items Detail", ""]
            ]
            
            # Add summary data
            for row_num, (label, value) in enumerate(summary_data, 1):
                ws_summary.cell(row=row_num, column=1, value=label).font = Font(bold=True)
                ws_summary.cell(row=row_num, column=2, value=value)
            
            # Add items detail
            start_row = len(summary_data) + 1
            item_headers = ["Item Name", "HSN/SAC", "Quantity", "Rate", "Item Total"]
            
            for col_num, header in enumerate(item_headers, 1):
                cell = ws_summary.cell(row=start_row, column=col_num, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="E6E6FA", end_color="E6E6FA", fill_type="solid")
            
            for row_num, item in enumerate(zoho_data['invoice_items'], start_row + 1):
                ws_summary.cell(row=row_num, column=1, value=item.get('Item Name', ''))
                ws_summary.cell(row=row_num, column=2, value=item.get('HSN/SAC', ''))
                ws_summary.cell(row=row_num, column=3, value=item.get('Quantity', ''))
                ws_summary.cell(row=row_num, column=4, value=f"₹{item.get('Rate', '0')}")
                ws_summary.cell(row=row_num, column=5, value=f"₹{item.get('Item Total', '0')}")
            
            # Auto-adjust column widths for summary sheet
            for column_cells in ws_summary.columns:
                length = max(len(str(cell.value or '')) for cell in column_cells)
                ws_summary.column_dimensions[column_cells[0].column_letter].width = min(length + 2, 50)
            
            # Make Zoho_Import_Format the active sheet
            wb.active = ws_zoho
            
            # Save the workbook
            wb.save(excel_filename)
            
            logger.info(f"Excel file created successfully: {excel_filename}")
            return excel_filename
            
        except Exception as e:
            logger.error(f"Error creating Excel file: {str(e)}")
            raise

    def process_invoice(self, pdf_path, export_excel=True):
        """Main function to process invoice and return Zoho-formatted data with Excel export"""
        try:
            logger.info(f"Processing invoice: {pdf_path}")
            
            with pdfplumber.open(pdf_path) as pdf:
                all_lines = []
                
                for page_num, page in enumerate(pdf.pages):
                    logger.info(f"Processing page {page_num + 1}")
                    text = page.extract_text(layout=True)
                    if text:
                        lines = text.split('\n')
                        all_lines.extend(lines)
            
            # Extract different sections
            metadata = self.extract_invoice_metadata(all_lines)
            items = self.extract_line_items(all_lines)
            tax_info = self.extract_tax_information(all_lines)
            totals = self.extract_totals(all_lines)
            
            logger.info(f"Extracted {len(items)} line items")
            logger.info(f"Found metadata: {list(metadata.keys())}")
            
            # Map to Zoho format
            zoho_data = self.map_to_zoho_format(metadata, items, tax_info, totals)
            
            # Export to Excel if requested
            result = zoho_data.copy()
            if export_excel:
                excel_filename = self.export_to_excel(zoho_data)
                result['excel_file'] = excel_filename
                result['excel_path'] = os.path.abspath(excel_filename)
            
            return result
            
        except Exception as e:
            logger.error(f"Error processing invoice: {str(e)}")
            raise

# Usage example
def main():
    """Example usage of the parser with Excel export"""
    parser = ZohoInvoiceParser()
    
    # Replace with your PDF path
    pdf_path = r"D:\Working_app_Invoice\TemplateMatching\invoice_extractor\data\invoice5_processed.pdf"
    
    try:
        # Process invoice and create Excel export
        result = parser.process_invoice(pdf_path, export_excel=True)
        
        # Display results
        print("=== NUCLEUS ANALYTICS INVOICE PROCESSED ===")
        print(f"Invoice Number: {result['invoice_header']['Bill Number']}")
        print(f"Invoice Date: {result['invoice_header']['Bill Date']}")
        print(f"Vendor: {result['invoice_header']['Vendor Name']}")
        print(f"Customer: {result['invoice_header']['Customer Name']}")
        print(f"GSTIN: {result['invoice_header']['GST Identification Number (GSTIN)']}")
        
        print(f"\n=== SUMMARY ===")
        print(f"Total Items: {result['summary']['total_items']}")
        print(f"Total Quantity: {result['summary']['total_quantity']}")
        print(f"Subtotal: ₹{result['summary']['subtotal']}")
        print(f"Total: ₹{result['summary']['total']}")
        
        if 'excel_file' in result:
            print(f"\n=== EXCEL EXPORT ===")
            print(f"Excel file created: {result['excel_file']}")
            print(f"Full path: {result['excel_path']}")
            print("\nThe Excel file contains two worksheets:")
            print("1. 'Zoho_Import_Format' - Ready for Zoho Books import")
            print("2. 'Invoice_Summary' - Human-readable summary")
        
        # Optional: Print sample of the first few columns for verification
        print(f"\n=== SAMPLE ZOHO FORMAT (First few fields) ===")
        if result['invoice_items']:
            sample_item = result['invoice_items'][0]
            header = result['invoice_header']
            sample_fields = ['Bill Date', 'Bill Number', 'Vendor Name', 'Item Name', 'Quantity', 'Rate', 'Item Total']
            
            for field in sample_fields:
                if field in header:
                    print(f"{field}: {header[field]}")
                elif field in sample_item:
                    print(f"{field}: {sample_item[field]}")
        
    except FileNotFoundError:
        print(f"Error: The file '{pdf_path}' was not found.")
        print("Please make sure the PDF file is in the same directory as the script.")
        print("You can also provide the full path to the PDF file.")
    except Exception as e:
        print(f"An error occurred: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    main()